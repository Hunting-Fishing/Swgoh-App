#!/usr/bin/env python3
"""Extract provenance-backed Guild Intelligence history from the Ludus workbook.

Requires Python 3 and openpyxl. This script is read-only: it never saves or mutates the XLSX.
It writes compact JSONL datasets consumed by scripts/import-ludus-workbook-history.mjs.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

SOURCE = "lv-unit-tracker-workbook"
EXPECTED_SHA256 = "4465e6e86933525be963ef10e93a12c4c33d861236f58d3a90864df91e57fe7d"
GL_IDS = {
    "GLAHSOKATANO", "JABBATHEHUTT", "JEDIMASTERKENOBI", "GRANDMASTERLUKE",
    "GLLEIA", "LORDVADER", "GLHONDO", "GLREY", "SITHPALPATINE",
    "SUPREMELEADERKYLOREN",
}
INQUISITOR_IDS = {
    "EIGHTHBROTHER", "FIFTHBROTHER", "GRANDINQUISITOR", "INQUISITORBARRISS",
    "MARROK", "NINTHSISTER", "SECONDSISTER", "SEVENTHSISTER", "THIRDSISTER",
}


def digits(value) -> str:
    return re.sub(r"\D", "", str(value or ""))


def integer(value, default=0):
    try:
        text = str(value).replace(",", "").strip()
        if text in {"", "--", "-", "None", "null"}:
            return default
        return int(float(text))
    except Exception:
        return default


def iso_utc(value):
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc).isoformat()
    return str(value or "")


def workbook_sha(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_jsonl(path: Path, rows):
    count = 0
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")
            count += 1
    return count


def member_snapshots(ws, sha):
    for index, row in enumerate(ws.iter_rows(values_only=True), 1):
        if index == 1 or len(row) < 4:
            continue
        try:
            captured_at = datetime.fromtimestamp(float(row[0]) / 1000.0, tz=timezone.utc).isoformat()
        except Exception:
            continue
        for raw in row[3:]:
            if not isinstance(raw, str) or not raw.startswith("{"):
                continue
            try:
                member = json.loads(raw)
            except Exception:
                continue
            ally_code = digits(member.get("ally_code"))
            if len(ally_code) != 9:
                continue
            units = {str(unit.get("base_id", "")): unit for unit in member.get("units", []) if isinstance(unit, dict)}
            gl_units = []
            for base_id in sorted(GL_IDS):
                unit = units.get(base_id)
                if unit and integer(unit.get("level")) > 0:
                    gl_units.append({
                        "baseId": base_id,
                        "gearLevel": integer(unit.get("gear_level")),
                        "level": integer(unit.get("level")),
                        "relicTier": integer(unit.get("relic_tier")),
                    })
            inquisitors = {}
            for base_id in sorted(INQUISITOR_IDS):
                unit = units.get(base_id)
                if unit:
                    inquisitors[base_id] = {
                        "gearLevel": integer(unit.get("gear_level")),
                        "level": integer(unit.get("level")),
                        "relicTier": integer(unit.get("relic_tier")),
                    }
            yield {
                "captured_at": captured_at,
                "ally_code": ally_code,
                "player_name": str(member.get("name", "")),
                "galactic_power": integer(member.get("galactic_power")),
                "character_power": integer(member.get("character_galactic_power")),
                "ship_power": integer(member.get("ship_galactic_power")),
                "guild_contribution": integer(member.get("guild_contribution")),
                "guild_exchange_donations": integer(member.get("guild_exchange_donations")),
                "gl_count": len(gl_units),
                "gl_units": gl_units,
                "inquisitor_units": inquisitors,
                "source": SOURCE,
                "source_ref": "Member Data_Backup_20260605_195",
                "metadata": {"workbookSha256": sha, "derivedFields": ["gl_count", "gl_units", "inquisitor_units"]},
            }


def raid_tickets(ws, sha):
    for index, row in enumerate(ws.iter_rows(values_only=True), 1):
        if index == 1 or len(row) < 4 or not row[3]:
            continue
        try:
            payload = json.loads(row[3])
        except Exception:
            continue
        date_value = row[1]
        ticket_date = date_value.date().isoformat() if isinstance(date_value, datetime) else str(date_value)[:10]
        for member in payload:
            ally_code = digits(member.get("ally_code"))
            if len(ally_code) != 9:
                continue
            yield {
                "captured_at": iso_utc(row[0]),
                "ticket_date": ticket_date,
                "ally_code": ally_code,
                "player_name": str(member.get("player_name", "")),
                "tickets": integer(member.get("tickets")),
                "source": SOURCE,
                "source_ref": "Raid Ticket Data",
                "metadata": {"workbookSha256": sha},
            }


def raid_results(ws, sha):
    for index, row in enumerate(ws.iter_rows(values_only=True), 1):
        if index == 1 or len(row) < 4 or not row[3]:
            continue
        try:
            payload = json.loads(row[3])
        except Exception:
            continue
        members = payload.get("memberData") or []
        scores = [integer(member.get("score")) for member in members]
        order = sorted(range(len(members)), key=lambda pos: scores[pos], reverse=True)
        ranks = {pos: rank + 1 for rank, pos in enumerate(order)}
        for pos, member in enumerate(members):
            ally_code = digits(member.get("ally_code"))
            yield {
                "raid_date": str(payload.get("date", ""))[:10],
                "raid_name": str(payload.get("raid", "")),
                "ally_code": ally_code if len(ally_code) == 9 else None,
                "player_name": str(member.get("player_name", "")),
                "score": scores[pos],
                "rank": ranks[pos],
                "source": SOURCE,
                "source_ref": "Endor Performance Data",
                "metadata": {"workbookSha256": sha, "rankDerived": True},
            }


def rote_performance(ws, sha):
    for index, row in enumerate(ws.iter_rows(values_only=True), 1):
        if index == 1 or len(row) < 3 or not row[2]:
            continue
        try:
            payload = json.loads(row[2])
        except Exception:
            continue
        start_date = str(payload.get("start_date", ""))
        for member in payload.get("performance_summary") or []:
            ally_code = digits(member.get("ally_code"))
            yield {
                "start_date": start_date,
                "ally_code": ally_code if len(ally_code) == 9 else None,
                "player_name": str(member.get("player_name", "")),
                "missed_phases": integer(member.get("missed_phases")),
                "missed_phase_one": integer(member.get("missed_phase_one")),
                "mission_attempts": integer(member.get("mission_attempts")),
                "missed_deployments": integer(member.get("missed_deployments")),
                "mission_tp": integer(member.get("mission_tp")),
                "deployed_tp": integer(member.get("deployed_tp")),
                "total_gp": integer(member.get("total_gp")),
                "zeffo": integer(member.get("zeffo")),
                "mandalore": integer(member.get("mandalore")),
                "reva": integer(member.get("reva")),
                "source": SOURCE,
                "source_ref": "ROTE Data",
                "metadata": {"workbookSha256": sha},
            }


def reva_history(ws, sha):
    for index, row in enumerate(ws.iter_rows(values_only=True), 1):
        if index == 1 or len(row) < 4 or not row[3]:
            continue
        try:
            payload = json.loads(row[3])
        except Exception:
            continue
        date_value = row[1]
        rote_start_date = date_value.date().isoformat() if isinstance(date_value, datetime) else str(date_value)[:10]
        earned = {}
        for member in payload.get("earned") or []:
            ally_code = digits(member.get("ally_code"))
            if len(ally_code) == 9:
                earned[ally_code] = str(member.get("player_name", ""))
        for raw_ally_code in payload.get("inGuild") or []:
            ally_code = digits(raw_ally_code)
            if len(ally_code) != 9:
                continue
            yield {
                "rote_start_date": rote_start_date,
                "ally_code": ally_code,
                "player_name": earned.get(ally_code) or None,
                "earned": ally_code in earned,
                "in_guild": True,
                "source": SOURCE,
                "source_ref": "ROTE Reva Shards",
                "metadata": {"workbookSha256": sha},
            }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out", type=Path, default=Path("data/ludus-history"))
    parser.add_argument("--allow-sha-mismatch", action="store_true")
    args = parser.parse_args()

    sha = workbook_sha(args.workbook)
    if sha != EXPECTED_SHA256 and not args.allow_sha_mismatch:
        raise SystemExit(f"Workbook SHA mismatch: expected {EXPECTED_SHA256}, got {sha}. Use --allow-sha-mismatch only for a reviewed replacement workbook.")

    args.out.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(args.workbook, read_only=True, data_only=False)
    datasets = {
        "member_snapshots.jsonl": member_snapshots(workbook["Member Data_Backup_20260605_195"], sha),
        "raid_tickets.jsonl": raid_tickets(workbook["Raid Ticket Data"], sha),
        "raid_results.jsonl": raid_results(workbook["Endor Performance Data"], sha),
        "rote_performance.jsonl": rote_performance(workbook["ROTE Data"], sha),
        "reva_shards.jsonl": reva_history(workbook["ROTE Reva Shards"], sha),
    }
    manifest = {"source": SOURCE, "workbookSha256": sha, "datasets": {}}
    for filename, rows in datasets.items():
        count = write_jsonl(args.out / filename, rows)
        manifest["datasets"][filename] = {"rows": count}
        print(f"{filename}: {count:,} rows")
    (args.out / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    print(f"manifest: {args.out / 'manifest.json'}")


if __name__ == "__main__":
    main()
